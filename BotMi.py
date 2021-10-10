import time
import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import Select

file = openpyxl.load_workbook('data\\userLogin.xlsx')
sheet = file.active
Datauser = []
Datapasss = []
barisMax = sheet.max_row

for i in range(2, barisMax + 1):
    userData = sheet.cell(row=i, column=1)
    Datauser.append(userData.value)

for i in range(2, barisMax + 1):
    userData = sheet.cell(row=i, column=2)
    Datapasss.append(userData.value)

def mov():
    alt = driver.window_handles
    for mv in alt:
        driver.switch_to.window(mv)
        print(driver.title)
        print(driver.current_window_handle)

def alamatTujuan():
    dataAlt = open('data\\alamatPengiriman.txt', 'r')
    nama = dataAlt.readline().strip()
    nm = driver.find_element_by_id('surName')
    nm.send_keys(nama)
    time.sleep(2)

    el = driver.find_element_by_id('UserAddressProvince')
    drp = Select(el)
    prov = dataAlt.readline().strip()
    all_opsi = drp.options
    for option in all_opsi:
        if option.text == prov:
            option.click()
            break
    time.sleep(2)
    el = driver.find_element_by_id('UserAddressCity')
    drp = Select(el)
    kota = dataAlt.readline().strip()
    all_opsi = drp.options
    for option in all_opsi:
        if option.text == kota:
            option.click()
            break
    time.sleep(2)
    kec = driver.find_element_by_id('UserAddressDistrict')
    drp = Select(kec)
    kecamatan = dataAlt.readline().strip()
    all_opsi = drp.options
    for option in all_opsi:
        if option.text == kecamatan:
            option.click()
            break
    time.sleep(2)
    kel = driver.find_element_by_id('UserAddressVillage')
    drp = Select(kel)
    kelurahan = dataAlt.readline().strip()
    all_opsi = drp.options
    for option in all_opsi:
        if option.text == kelurahan:
            option.click()
            break
    time.sleep(2)
    alamat = dataAlt.readline().strip()
    alt = driver.find_element_by_id('userAddress1')
    alt.send_keys(alamat)
    time.sleep(5)
    email = dataAlt.readline().strip()
    eml = driver.find_element_by_id('email')
    eml.send_keys(email)
    time.sleep(5)
    telpon = dataAlt.readline().strip()
    tlp = driver.find_element_by_id('Telephone')
    tlp.send_keys(telpon)
    dataAlt.close()

    time.sleep(5)
    byr = driver.find_element_by_xpath('/html/body/div[6]/div/div[2]/button[2]').click()
    time.sleep(15)
    lanjut = driver.find_element_by_xpath('/html/body/div[5]/div[2]/input[1]').click()
    time.sleep(3)

    driver.find_element_by_xpath('/html/body/div[1]/header/div/div/div[1]/a').click()
    time.sleep(3)
    driver.close()

goOrder = True
hop = barisMax - 1
ke = 1
while goOrder:
    try:
        print("login Akun ke:", ke)
        user = Datauser[ke-1]
        passs = Datapasss[ke-1]
        driver = webdriver.Firefox()
        driver.get('https://buy.mi.co.id/id/site/login')
        time.sleep(3)
        loginUser = driver.find_element_by_id('username')
        loginUser.send_keys(user)
        print("[+] Email: "+user)
        loginPass = driver.find_element_by_id('pwd')
        loginPass.send_keys(passs)
        print("[+] Password: ********")
        letsGo = driver.find_element_by_id('login-button').click()

        ads = driver.find_element_by_xpath("/html/body/div[4]/div/a")
        if ads:
            print ("Ads Found")
            cls = driver.find_element_by_xpath("/html/body/div[4]/div/div/span").click()
        else:
            print("Ads Not Found")

        time.sleep(10)
        ap = driver.find_element_by_xpath('/html/body/nav/div/div[1]/div[2]/div/dl[7]/dt/a').click()
        pilih = driver.find_element_by_xpath('/html/body/div[4]/div[1]/div[2]/ul/li[10]').click()
        beli = driver.find_element_by_xpath('/html/body/div[3]/div/div/ul/li[4]/a').click()
        mov()
        time.sleep(5)
        keranjang = driver.find_element_by_xpath('/html/body/div[1]/main/div/article/section[8]/div[1]/div[1]/button').click()
        time.sleep(5)
        byr = driver.find_element_by_xpath('/html/body/div[1]/main/footer/div/button[2]').click()
        add = driver.find_element_by_xpath('/html/body/div[4]/div[2]/div/div[2]/div').click()
        alamatTujuan()

        mov()
        time.sleep(5)
        us = driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[2]/div[1]")
        akun = driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[2]/div[1]/div/ul/li[1]/a")
        ulasan = driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[2]/div[1]/div/ul/li[2]/a")
        logout = driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div[2]/div[1]/div/ul/li[3]/a")
        action = ActionChains(driver)
        action.move_to_element(us).move_to_element(akun).move_to_element(ulasan).move_to_element(logout).click().perform()
        time.sleep(2)
        driver.close()

    except IndexError:
        pass
    if ke == hop:
        goOrder = False
    ke += 1
    continue


